"""
Author: Motty Mathews
Linkedin: www.linkedin.com/in/motty-mathews-35b0b882
Date: 2025-07-28
Description: The script reads an Excel file containing family tree data and converts it into a GEDCOM format.
Usage: python familyTree_excel_to_gedcom_final.py
Dependencies: openpyxl, pandas, dateutil
Note: Ensure the Excel file is structured correctly with the expected headers.
This script includes validation checks for the Excel data, ensuring no self-references and no duplicate full names.
It generates a validation log if issues are found.
"""


from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict
import logging
from dateutil import parser
import os

logging.basicConfig(level=logging.INFO, format='%(message)s')

def normalize_name(name):
    return ' '.join(name.strip().split()) if name else ''

def fmt_date(date_obj):
    if isinstance(date_obj, datetime):
        return date_obj.strftime('%d %b %Y').upper()
    return None

def try_parse_date(value):
    if isinstance(value, datetime):
        return value
    try:
        return parser.parse(str(value), fuzzy=True)
    except:
        return None

def excel_to_gedcom(input_path, output_path):
    wb = load_workbook(input_path)
    sheet = wb["RealWorkingSheet"]

    headers = [str(cell.value).strip() for cell in sheet[1]]
    logging.info(f"🔍 Excel column headers: {headers}")
    people_data = []

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        person = dict(zip(headers, row))
        logging.info(f"📄 Row {row_index}: {person}")

        person['Full Name'] = normalize_name(person.get('Full Name'))
        person['Father Full name'] = normalize_name(person.get('Father Full name'))
        person['Mother Full Name'] = normalize_name(person.get('Mother Full Name'))
        person['Spouse Full Name'] = normalize_name(person.get('Spouse Full Name'))
        person['Children'] = [normalize_name(c) for c in str(person.get('Children') or '').split(',') if c.strip()]
        people_data.append(person)

    all_names = set()
    for p in people_data:
        all_names.add(p['Full Name'])
        all_names.update([p['Father Full name'], p['Mother Full Name'], p['Spouse Full Name']])
        all_names.update(p['Children'])
    all_names = sorted(n for n in all_names if n)
    id_map = {name: f"@I{i + 1}@" for i, name in enumerate(all_names)}

    couple_families = {}
    fam_list = []
    fam_index = 1
    famc = defaultdict(list)
    fams = defaultdict(list)

    for p in people_data:
        pname = p['Full Name']
        father = p['Father Full name']
        mother = p['Mother Full Name']
        spouse = p['Spouse Full Name']
        children = p['Children']      

        if father or mother:
            key = (father, mother)
            if key not in couple_families:
                fam_id = f"@F{fam_index}@"
                fam_index += 1
                couple_families[key] = {'id': fam_id, 'husband': father, 'wife': mother, 'children': []}
                fam_list.append(couple_families[key])
            couple_families[key]['children'].append(pname)
            famc[pname].append(couple_families[key]['id'])
            if father: fams[father].append(couple_families[key]['id'])
            if mother: fams[mother].append(couple_families[key]['id'])

        if spouse:
            gender = (p.get('Gender') or '').strip().lower()
            if gender in ['m', 'male']:
                husband, wife = pname, spouse
            elif gender in ['f', 'female']:
                husband, wife = spouse, pname
            else:
                husband, wife = pname, spouse

            key = (husband, wife)
            if key not in couple_families:
                fam_id = f"@F{fam_index}@"
                fam_index += 1
                couple_families[key] = {'id': fam_id, 'husband': husband, 'wife': wife, 'children': []}
                fam_list.append(couple_families[key])
            couple_families[key]['children'].extend(children)
            fams[husband].append(couple_families[key]['id'])
            fams[wife].append(couple_families[key]['id'])
            for child in children:
                famc[child].append(couple_families[key]['id'])

    for fam in fam_list:
        fam['children'] = sorted(set(fam['children']))

    ged = ["0 HEAD", "1 SOUR CleanExcel2GED", "1 GEDC", "2 VERS 5.5.1", "1 CHAR UTF-8"]

    for name in all_names:
        iid = id_map[name]
        ged.append(f"0 {iid} INDI")
        parts = name.split()
        given = ' '.join(parts[:-1]) if len(parts) > 1 else name
        surname = parts[-1] if len(parts) > 1 else ''
        ged.append(f"1 NAME {given} /{surname}/")

        row = next((x for x in people_data if normalize_name(x['Full Name']) == normalize_name(name)), None)
        if row:
            gender = (row.get('Gender') or '').strip().lower()
            if gender in ['m', 'male']:
                ged.append("1 SEX M")
            elif gender in ['f', 'female']:
                ged.append("1 SEX F")

            dob_raw = row.get('Date of Birth (mm/dd/yyyy)') or row.get('DOB')
            dod_raw = row.get('Date of Death (mm/dd/yyyy)')
            dob = try_parse_date(dob_raw)
            dod = try_parse_date(dod_raw)

            bdate = fmt_date(dob)
            if bdate:
                ged.append("1 BIRT")
                ged.append(f"2 DATE {bdate}")
                logging.info(f"✅ DOB added for: {name} → {bdate}")
            else:
                logging.info(f"⚠️ No DOB found for: {name}")

            ddate = fmt_date(dod)
            if ddate:
                ged.append("1 DEAT")
                ged.append(f"2 DATE {ddate}")
                logging.info(f"✅ DOD added for: {name} → {ddate}")
            else:
                status = str(row.get('Dead/Alive', '')).strip().lower()
                if status == 'alive':
                    ged.append("1 ALIV Y")
                    logging.info(f"✅ Marked alive: {name}")
                elif status == 'dead':
                    ged.append("1 DEAT")
                    logging.info(f"⚰️ Marked dead (no date): {name}")
                else:
                    logging.info(f"⚠️ No DOD and no status: {name}")

            phone = str(row.get('Phone Number') or '').strip()
            addr1 = str(row.get('Address Line 1') or '').strip()
            addr2 = str(row.get('Address Line 2') or '').strip()
            city = str(row.get('City') or '').strip()
            state = str(row.get('State') or '').strip()
            zipc = str(row.get('Zipcode') or '').strip()
            country = str(row.get('Country') or '').strip()

            if phone:
                ged.append(f"1 PHON {phone}")
            if addr1 or addr2 or city or state or zipc or country:
                ged.append("1 ADDR " + addr1)
                if addr2:
                    ged.append(f"2 CONT {addr2}")
                if city:
                    ged.append(f"2 CITY {city}")
                if state:
                    ged.append(f"2 STAE {state}")
                if zipc:
                    ged.append(f"2 POST {zipc}")
                if country:
                    ged.append(f"2 CTRY {country}")


        for f in sorted(set(famc[name])):
            ged.append(f"1 FAMC {f}")
        for f in sorted(set(fams[name])):
            ged.append(f"1 FAMS {f}")

    for fam in fam_list:
        ged.append(f"0 {fam['id']} FAM")
        if fam['husband'] in id_map:
            ged.append(f"1 HUSB {id_map[fam['husband']]}")
        if fam['wife'] in id_map:
            ged.append(f"1 WIFE {id_map[fam['wife']]}")
        for ch in fam['children']:
            if ch in id_map:
                ged.append(f"1 CHIL {id_map[ch]}")

    ged.append("0 TRLR")

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(ged))
    logging.info(f"✅ GEDCOM file saved to: {output_path}")


import pandas as pd
from collections import defaultdict

def normalize(name):
    return ' '.join(str(name).strip().lower().split()) if pd.notna(name) else None

def validate_excel(path, sheetname="RealWorkingSheet", log_path="family-tree-validation-log.txt"):
    if os.path.exists(log_path):
        os.remove(log_path)
    result = {"log_path": log_path, "has_issues": False, "messages": []}
    result = {"log_path": log_path, "has_issues": False, "messages": []}
    df = pd.read_excel(path, sheet_name=sheetname)

    declared_names = set(df['Full Name'].map(normalize).dropna())

    referenced_names = set()
    for col in ['Father Full name', 'Mother Full Name', 'Spouse Full Name']:
        referenced_names.update(df[col].map(normalize).dropna())

    for children in df['Children'].dropna():
        referenced_names.update([normalize(name) for name in str(children).split(',') if name.strip()])

    referenced_only = []  # ignored
    duplicate_full_names = df['Full Name'].map(normalize).value_counts()
    duplicates = duplicate_full_names[duplicate_full_names > 1].to_dict()

    invalid_self_refs = []
    for idx, row in df.iterrows():
        full_name = normalize(row.get('Full Name'))
        if not full_name:
            continue
        matches = []
        for rel in ['Father Full name', 'Mother Full Name', 'Spouse Full Name']:
            if normalize(row.get(rel)) == full_name:
                matches.append(rel)
        children = [normalize(c) for c in str(row.get('Children') or '').split(',') if c.strip()]
        if full_name in children:
            matches.append('Children')
        if matches:
            invalid_self_refs.append((idx + 2, row.get('Full Name'), matches))

    has_issues = referenced_only or duplicates or invalid_self_refs
    if has_issues:
        result["has_issues"] = True
        with open(log_path, "w", encoding="utf-8") as log:
            log.write("DATA VALIDATION SUMMARY:\n\n")
            if duplicates:
                log.write("❗ Duplicate Full Names:\n")
                for name, count in duplicates.items():
                    log.write(f" - {name} (count: {count})\n")
                log.write("\n")
            if referenced_only:
                log.write("⚠️ Referenced but Undeclared Names:\n")
                for name in referenced_only:
                    log.write(f" - {name}\n")
                log.write("\n")
            if invalid_self_refs:
                log.write("🛑 Self-References Found:\n")
                for rownum, name, fields in invalid_self_refs:
                    log.write(f" - Row {rownum}: '{name}' listed in {', '.join(fields)}\n")
                result["messages"].append(f"⚠️ Data issues found. See validation log: {log_path}")
    else:
            result["messages"].append("✅ No data validation issues found.")

    result["duplicates"] = list(duplicates.keys())
    result["self_references"] = [entry[1] for entry in invalid_self_refs]
    return result



if __name__ == "__main__":
    
    excel_path = "family_tree_sample.xlsx"
    validation_result = validate_excel(excel_path)
    base_name = os.path.splitext(os.path.basename(excel_path))[0]
    gedcom_path = f"{base_name}.ged"
    excel_to_gedcom(excel_path, gedcom_path)

    for line in validation_result["messages"]:
        print(line)

    if validation_result["duplicates"]:
        print("\n❗ Duplicate Full Names:")
        for name in validation_result["duplicates"]:
            print(f" - {name}")

    if validation_result["self_references"]:
        print("\n🛑 Self-referencing entries:")
        for name in validation_result["self_references"]:
            print(f" - {name}")

        

